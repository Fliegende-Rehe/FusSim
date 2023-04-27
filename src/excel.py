import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from .fusion import *


class ExcelWorkbook:
    def __init__(self):
        self.workbook = None
        self.column_names = {}
        self.parent_column_names = {}

    def createWorkbook(self, path):
        self.workbook = Workbook()
        self.saveWorkbook(path)

    def saveWorkbook(self, path):
        self.workbook.save(path)

    def openWorkbook(self, path):
        if os.path.isfile(path):
            self.workbook = load_workbook(filename=path)
        else:
            self.createWorkbook(path)
        fusion.logger('Open Time Study Table')

    def writeToCell(self, columnName, row, value):
        column = self.get_column_index(columnName)
        cell = self.workbook.active.cell(row=row, column=column)
        cell.value = value

    def readCell(self, columnName, row):
        column = self.get_column_index(columnName)
        cell = self.workbook.active.cell(row=row, column=column)
        return cell.value

    def setColumnName(self, column, name):
        self.column_names[column_index_from_string(column)] = name

    def get_column_name(self, column):
        if column in self.column_names:
            return self.column_names[column]
        else:
            return get_column_letter(column)

    def get_column_index(self, column_name):
        return column_index_from_string(column_name)
